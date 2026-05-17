"""Generowanie raportów CSV / XLSX z danych analityki."""
import csv
import io

from . import metrics


def _report_rows(report):
    """Zwraca (nagłówki, wiersze) dla danego raportu."""
    if report == 'overview':
        d = metrics.overview()
        sb = d['status_breakdown']
        headers = ['Metryka', 'Wartość']
        rows = [
            ['Wszystkie pomysły', d['total_ideas']],
            ['Do weryfikacji', sb.get('TO_VERIFY', 0)],
            ['Zgłoszone', sb.get('SUBMITTED', 0)],
            ['W trakcie', sb.get('IN_PROGRESS', 0)],
            ['Wdrożone', sb.get('IMPLEMENTED', 0)],
            ['Odrzucone', sb.get('CANCELLED', 0)],
            ['% wdrożeń', d['implemented_rate']],
            ['Śr. czas akceptacji (h)', d['avg_approval_hours']],
            ['Śr. postęp wdrożeń (%)', d['avg_progress']],
            ['Oszczędności zrealizowane (zł)', d['savings']['realized_money']],
            ['Oszczędności potencjalne (zł)', d['savings']['potential_money']],
            ['Oszczędności czasu (h)', d['savings']['realized_hours']],
        ]
        return headers, rows

    if report == 'departments':
        data = metrics.departments()
        headers = ['Dział', 'Pomysły', 'Wdrożone', '% wdrożeń',
                   'Oszczędności (zł)', 'Koszt (zł)', 'ROI', 'Śr. akceptacja (h)']
        rows = [
            [r['department'], r['total_ideas'], r['implemented'],
             r['implemented_rate'], r['savings_money'], r['estimated_cost'],
             r['roi'], r['avg_approval_hours']]
            for r in data
        ]
        return headers, rows

    if report == 'categories':
        data = metrics.categories()
        headers = ['Kategoria', 'Pomysły', 'Wdrożone', '% wdrożeń', 'Oszczędności (zł)']
        rows = [
            [r['category'], r['total_ideas'], r['implemented'],
             r['implemented_rate'], r['savings_money']]
            for r in data
        ]
        return headers, rows

    if report == 'trends':
        data = metrics.trends()
        headers = ['Okres', 'Zgłoszenia', 'Wdrożenia', 'Oszczędności (zł)']
        rows = [
            [r['period'], r['submissions'], r['implementations'], r['savings']]
            for r in data
        ]
        return headers, rows

    raise ValueError(f'Nieznany raport: {report}')


def to_csv(report):
    headers, rows = _report_rows(report)
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=';')
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode('utf-8-sig')


def to_xlsx(report):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    headers, rows = _report_rows(report)
    wb = Workbook()
    ws = wb.active
    ws.title = report[:31]

    header_fill = PatternFill('solid', fgColor='1D2B64')
    header_font = Font(bold=True, color='FFFFFF')
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        ws.append(row)

    for col_idx, _ in enumerate(headers, start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        width = max(
            len(str(headers[col_idx - 1])),
            *(len(str(r[col_idx - 1])) for r in rows) if rows else [10],
        )
        ws.column_dimensions[letter].width = min(40, width + 4)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
